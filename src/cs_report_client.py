"""Client for the Customer Success Report exported to Google Drive.

Reads the latest CS Report spreadsheet from the 'Data Exports' shared drive,
parses the JSON-encoded KPI values, and provides per-customer summaries for
platform health, supply chain metrics, and ROI/value metrics.

Cross-validates overlapping fields with Pendo data via the QA registry.
"""

from __future__ import annotations

import io
import json
import threading
from datetime import date, datetime
from collections import defaultdict
from pathlib import Path
from typing import Any

import yaml

from .config import logger
from .config_paths import COHORTS_FILE, CS_REPORT_CUSTOMER_ALIASES_FILE
from .qa import qa

# Shared Drive ID and folder for the CS Report
_DATA_EXPORTS_DRIVE_ID = "0AHL7kClvi-JmUk9PVA"
_CS_REPORT_FOLDER_ID = "16922c1MWTKNx3Pw1W7bbUARC_q2aat0Z"

_cache: dict[str, Any] | None = None
_cache_lock = threading.Lock()

# Optional: project-root YAML — map Pendo customer → exact CS Report `customer` values
_CSR_ALIAS_FILE = CS_REPORT_CUSTOMER_ALIASES_FILE
_COHORTS_FILE = COHORTS_FILE
_cs_report_alias_map: dict[str, list[str]] | None = None
_cs_report_alias_lock = threading.Lock()
_cohort_customer_alias_map: dict[str, list[str]] | None = None
_cohort_customer_alias_lock = threading.Lock()


def _get_drive():
    from .slides_api import _get_service
    _x, drive, _sh = _get_service()
    return drive


def _parse_kpi(raw) -> dict[str, Any] | None:
    """Parse a JSON-encoded KPI cell into a dict with startValue/endValue/delta."""
    if raw is None:
        return None
    try:
        d = json.loads(str(raw))
        if isinstance(d, dict) and not d.get("empty", False):
            return d
    except (json.JSONDecodeError, TypeError):
        pass
    return None


def _kpi_end(raw) -> float | None:
    """Extract the end-of-period value from a KPI cell."""
    d = _parse_kpi(raw)
    if d is None:
        return None
    v = d.get("endValue")
    if v is None:
        v = d.get("startValue")
    return float(v) if v is not None else None


_HEALTH_SCORE_COLORS = frozenset({"GREEN", "YELLOW", "RED", "NONE"})


def _health_bucket_from_numeric(score: float) -> str:
    """Map CSR automated composite 0–100 to display bucket (when export column is NONE)."""
    if score >= 80.0:
        return "GREEN"
    if score >= 60.0:
        return "YELLOW"
    return "RED"


def _health_bucket_from_automated_row(row: dict[str, Any]) -> str | None:
    """Read ``automatedHealthScores`` JSON when the export ``healthScore`` cell is NONE."""
    raw = row.get("automatedHealthScores")
    if not raw or not isinstance(raw, str) or not raw.strip().startswith("["):
        return None
    try:
        payload = json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return None
    if not isinstance(payload, list) or not payload:
        return None
    item = payload[0]
    if not isinstance(item, dict):
        return None
    override = item.get("override")
    if isinstance(override, str):
        ov = override.strip().upper()
        if ov in _HEALTH_SCORE_COLORS and ov != "NONE":
            return ov
    composite = item.get("healthScore")
    if isinstance(composite, (int, float)):
        return _health_bucket_from_numeric(float(composite))
    return None


def _normalize_health_score(raw: Any) -> str:
    """CSR health bucket (GREEN/YELLOW/RED/NONE); accepts plain strings or JSON KPI cells."""
    if raw is None:
        return "NONE"
    if isinstance(raw, (int, float)):
        return "NONE"
    if isinstance(raw, str):
        s = raw.strip()
        if not s:
            return "NONE"
        if s.startswith("{"):
            d = _parse_kpi(s)
            if d:
                v = d.get("endValue")
                if v is None:
                    v = d.get("startValue")
                if v is not None:
                    s = str(v).strip()
        upper = s.upper()
        if upper in ("GREEN", "YELLOW", "RED", "NONE"):
            return upper
        return s
    return "NONE"


def _health_score_from_row(row: dict[str, Any]) -> str:
    """Resolve site health: ``healthScore`` column, else ``automatedHealthScores`` composite."""
    column_val: Any = None
    for col in ("healthScore", "health_score", "Health Score"):
        if col in row:
            column_val = row.get(col)
            break
    if column_val is None:
        for k, v in row.items():
            if str(k).strip().lower() == "healthscore":
                column_val = v
                break
    if column_val is not None:
        bucket = _normalize_health_score(column_val)
        if bucket != "NONE":
            return bucket
    fallback = _health_bucket_from_automated_row(row)
    if fallback:
        return fallback
    if column_val is not None:
        return _normalize_health_score(column_val)
    return "NONE"


def _csr_row_dedupe_key(row: dict[str, Any]) -> str:
    parts = [
        (row.get("customer") or "").strip().lower(),
        (row.get("factoryName") or "").strip().lower(),
        (row.get("entity") or row.get("Entity") or "").strip().lower(),
        (row.get("site") or row.get("Site") or "").strip().lower(),
    ]
    return "|".join(parts)


def _kpi_delta_pct(raw) -> float | None:
    """Extract the delta percentage from a KPI cell."""
    d = _parse_kpi(raw)
    if d is None:
        return None
    v = d.get("deltaPercent")
    return float(v) if v is not None else None


def check_reachable() -> None:
    """Verify the CS Report folder on Drive is reachable. Raises if Drive or folder is down."""
    from .network_utils import network_timeout
    
    with network_timeout(30.0, "Drive CS Report folder check"):
        drive = _get_drive()
        q = f"'{_CS_REPORT_FOLDER_ID}' in parents and trashed = false"
        drive.files().list(
            q=q,
            fields="files(id, name)",
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            corpora="drive",
            driveId=_DATA_EXPORTS_DRIVE_ID,
            pageSize=1,
        ).execute()


def _fetch_latest_report() -> list[dict[str, Any]]:
    """Download the latest CS Report from Drive and parse all rows.

    Thread-safe: a lock ensures only one thread fetches/parses the XLSX while
    others wait for the cached result.
    """
    global _cache
    if _cache is not None:
        return _cache["rows"]

    with _cache_lock:
        # Double-check after acquiring lock (another thread may have populated it)
        if _cache is not None:
            return _cache["rows"]

        from .network_utils import network_timeout
        with network_timeout(30.0, "Drive CS Report download"):
            drive = _get_drive()

            q = f"'{_CS_REPORT_FOLDER_ID}' in parents and trashed = false"
            results = drive.files().list(
                q=q,
                fields="files(id, name, modifiedTime)",
                includeItemsFromAllDrives=True,
                supportsAllDrives=True,
                corpora="drive",
                driveId=_DATA_EXPORTS_DRIVE_ID,
                pageSize=5,
                orderBy="modifiedTime desc",
            ).execute()
            files = results.get("files", [])
            if not files:
                logger.warning("No CS Report found in Data Exports drive")
                return []

            latest = files[0]
            logger.info("Fetching CS Report: %s (%s)", latest["name"], latest["modifiedTime"][:10])

            request = drive.files().export_media(
                fileId=latest["id"],
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            buf = io.BytesIO()
            from googleapiclient.http import MediaIoBaseDownload
            downloader = MediaIoBaseDownload(buf, request)
            done = False
            chunk_count = 0
            while not done:
                _, done = downloader.next_chunk()
                chunk_count += 1
                if chunk_count > 100:  # Safety limit: max 100 chunks
                    raise TimeoutError(f"CS Report download exceeded max chunks (100)")

        buf.seek(0)
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError(
                "CS Report XLSX parsing requires openpyxl; add it to your environment "
                "(e.g. pip install openpyxl or pip install -r requirements.txt)."
            ) from e
        wb = openpyxl.load_workbook(buf, read_only=True)
        ws = wb[wb.sheetnames[0]]

        headers: list[str] = []
        rows: list[dict[str, Any]] = []
        for ri, row in enumerate(ws.iter_rows()):
            cells = list(row)
            if ri == 0:
                headers = [str(c.value) if c.value else "" for c in cells]
                continue
            vals: dict[str, Any] = {}
            for i, c in enumerate(cells):
                if i < len(headers) and headers[i]:
                    vals[headers[i]] = c.value
            if vals.get("customer"):
                rows.append(vals)
        wb.close()

        _cache = {"rows": rows, "file": latest["name"], "modified": latest["modifiedTime"]}
        logger.info("Loaded %d rows from CS Report (%d columns)", len(rows), len(headers))
        return rows


def _load_cs_report_alias_map() -> dict[str, list[str]]:
    """Pendo (or QBR) customer name (lower) → list of exact CS `customer` column values."""
    global _cs_report_alias_map
    if _cs_report_alias_map is not None:
        return _cs_report_alias_map
    with _cs_report_alias_lock:
        if _cs_report_alias_map is not None:
            return _cs_report_alias_map
        out: dict[str, list[str]] = {}
        if _CSR_ALIAS_FILE.is_file():
            try:
                raw = yaml.safe_load(_CSR_ALIAS_FILE.read_text())
                if isinstance(raw, dict):
                    for k, v in raw.items():
                        if str(k).strip().startswith("#"):
                            continue
                        key = str(k).strip().lower()
                        if not key:
                            continue
                        if isinstance(v, str):
                            vals = [v]
                        elif isinstance(v, list):
                            vals = [str(x).strip() for x in v if str(x).strip()]
                        else:
                            continue
                        out[key] = vals
            except Exception as e:
                logger.warning("cs_report_customer_aliases: could not load %s: %s", _CSR_ALIAS_FILE, e)
        _cs_report_alias_map = out
        return out


def _load_cohort_customer_alias_map() -> dict[str, list[str]]:
    """Map cohort key / name / alias → all related account labels (for CSR lookup expansion)."""
    global _cohort_customer_alias_map
    if _cohort_customer_alias_map is not None:
        return _cohort_customer_alias_map
    with _cohort_customer_alias_lock:
        if _cohort_customer_alias_map is not None:
            return _cohort_customer_alias_map
        out: dict[str, list[str]] = {}
        if _COHORTS_FILE.is_file():
            try:
                data = yaml.safe_load(_COHORTS_FILE.read_text()) or {}
            except Exception as e:
                logger.warning("cohorts aliases (CS Report): could not load %s: %s", _COHORTS_FILE, e)
                data = {}
            customers = data.get("customers") if isinstance(data, dict) else None
            if not isinstance(customers, dict) and isinstance(data, dict):
                customers = data.get("cohorts")
            if isinstance(customers, dict):
                for key, row in customers.items():
                    terms: list[str] = [str(key).strip()]
                    if isinstance(row, dict):
                        name = str(row.get("name") or "").strip()
                        if name:
                            terms.append(name)
                        aliases = row.get("aliases") or []
                        if isinstance(aliases, str):
                            aliases = [aliases]
                        if isinstance(aliases, (list, tuple)):
                            terms.extend(str(a).strip() for a in aliases if str(a).strip())
                    deduped: list[str] = []
                    seen: set[str] = set()
                    for t in terms:
                        if t and t.lower() not in seen:
                            seen.add(t.lower())
                            deduped.append(t)
                    for t in deduped:
                        out[t.lower()] = deduped
        _cohort_customer_alias_map = out
        return out


def cs_report_lookup_keys_for_account(
    *,
    salesforce_label: str = "",
    pendo_customer_key: str | None = None,
) -> list[str]:
    """Ordered lookup keys for CS Report row matching (SF label, Pendo prefix, cohort aliases)."""
    keys: list[str] = []

    def add(term: str) -> None:
        s = (term or "").strip()
        if not s:
            return
        if s.lower() in {k.lower() for k in keys}:
            return
        keys.append(s)

    add(salesforce_label)
    add(pendo_customer_key or "")
    csr = _load_cs_report_alias_map()
    for seed in list(keys):
        for term in csr.get(seed.lower(), []):
            add(term)
    cohort = _load_cohort_customer_alias_map()
    for seed in list(keys):
        for term in cohort.get(seed.lower(), []):
            add(term)
    return keys


def cs_report_customer_name_candidates(pendo_name: str) -> list[str]:
    """Return distinct names to try for CS `customer` matching: pendo name first, then aliases."""
    raw = (pendo_name or "").strip()
    if not raw:
        return []
    al = _load_cs_report_alias_map().get(raw.lower()) or []
    out: list[str] = []
    seen: set[str] = set()
    for c in (raw, *al):
        s = (c or "").strip()
        if not s:
            continue
        k = s.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(s)
    return out


def _customer_rows(customer_name: str, delta: str = "week") -> list[dict[str, Any]]:
    """Get rows for one lookup key (plus ``config/cs_report_customer_aliases.yaml`` candidates)."""
    sites, _matched, _tried, _merged = _sites_for_customer_lookup(customer_name, delta=delta)
    return sites


def _sites_for_customer_lookup(
    primary_name: str,
    *,
    lookup_keys: list[str] | None = None,
    delta: str = "week",
) -> tuple[list[dict[str, Any]], str | None, list[str], list[str]]:
    """Merge week rows for every CSR ``customer`` resolved from *lookup_keys* and aliases.

    Returns ``(rows, matched_lookup_key, tried_customer_values, matched_csr_customers)``.
    """
    rows = _fetch_latest_report()
    keys: list[str] = []
    for k in lookup_keys or []:
        s = (k or "").strip()
        if s and s.lower() not in {x.lower() for x in keys}:
            keys.append(s)
    primary = (primary_name or "").strip()
    if primary and primary.lower() not in {x.lower() for x in keys}:
        keys.insert(0, primary)
    if not keys and primary:
        keys = [primary]

    all_tried: list[str] = []
    seen_tried: set[str] = set()
    merged: list[dict[str, Any]] = []
    seen_row_keys: set[str] = set()
    matched_csr_customers: list[str] = []
    matched_lookup_key: str | None = None

    for key in keys:
        cands = cs_report_customer_name_candidates(key)
        key_lower = key.lower()
        for name in cands:
            nl = name.lower()
            if nl in seen_tried:
                continue
            seen_tried.add(nl)
            all_tried.append(name)
            matched = [
                r
                for r in rows
                if (r.get("customer") or "").strip().lower() == nl and r.get("delta") == delta
            ]
            if not matched:
                continue
            if matched_lookup_key is None:
                matched_lookup_key = key
            if name not in matched_csr_customers:
                matched_csr_customers.append(name)
            if nl != key_lower:
                logger.info(
                    "CS Report: matched %d row(s) for lookup key %r using `customer`=%r",
                    len(matched),
                    key,
                    name,
                )
            for r in matched:
                rk = _csr_row_dedupe_key(r)
                if rk in seen_row_keys:
                    continue
                seen_row_keys.add(rk)
                merged.append(r)

    if merged and len(matched_csr_customers) > 1:
        logger.info(
            "CS Report: merged %d row(s) across CSR customer values %r for account %r",
            len(merged),
            matched_csr_customers,
            primary_name or matched_lookup_key,
        )
    return merged, matched_lookup_key, all_tried, matched_csr_customers


def _add_site_entity_from_row(row: dict[str, Any], entry: dict[str, Any]) -> None:
    """If the CS Report row has site/entity columns (any casing), add them to entry."""
    for key, val in row.items():
        if val is None or val == "":
            continue
        k = key.strip().lower()
        if k == "site":
            entry["site"] = val if isinstance(val, str) else str(val)
        elif k == "entity":
            entry["entity"] = val if isinstance(val, str) else str(val)


def _set_csr_kpi_int(entry: dict[str, Any], row: dict[str, Any], column: str, export_key: str) -> None:
    val = _kpi_end(row.get(column))
    if val is not None:
        entry[export_key] = int(val)


def _set_csr_kpi_dec(
    entry: dict[str, Any],
    row: dict[str, Any],
    column: str,
    export_key: str,
    *,
    decimals: int | None,
) -> None:
    val = _kpi_end(row.get(column))
    if val is None:
        return
    if decimals is None:
        entry[export_key] = round(val)
    else:
        entry[export_key] = round(val, decimals)


def _csr_plain_scalar(raw: Any) -> Any | None:
    """Normalize CS Report plain column values for export (JSON/Sheets-safe)."""
    if raw is None or raw == "":
        return None
    if isinstance(raw, str):
        text = raw.strip()
        return text if text else None
    if isinstance(raw, datetime):
        if raw.tzinfo is not None:
            return raw.isoformat()
        if raw.hour == 0 and raw.minute == 0 and raw.second == 0 and raw.microsecond == 0:
            return raw.date().isoformat()
        return raw.isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    return raw


def _set_csr_plain(entry: dict[str, Any], row: dict[str, Any], column: str, export_key: str) -> None:
    scalar = _csr_plain_scalar(row.get(column))
    if scalar is not None:
        entry[export_key] = scalar


def _add_automated_health_export(row: dict[str, Any], entry: dict[str, Any]) -> None:
    raw = row.get("automatedHealthScores")
    if raw is None or raw == "":
        return
    if isinstance(raw, str):
        entry["automated_health_scores"] = raw.strip()
    else:
        entry["automated_health_scores"] = json.dumps(raw, default=str)
    payload = None
    if isinstance(raw, str) and raw.strip().startswith("["):
        try:
            payload = json.loads(raw)
        except (json.JSONDecodeError, TypeError, ValueError):
            payload = None
    elif isinstance(raw, list):
        payload = raw
    if isinstance(payload, list) and payload:
        item = payload[0]
        if isinstance(item, dict):
            composite = item.get("healthScore")
            if isinstance(composite, (int, float)):
                entry["automated_health_composite"] = round(float(composite), 1)
            override = item.get("override")
            if isinstance(override, str) and override.strip():
                entry["automated_health_override"] = override.strip()


# KPI columns → export keys. ``decimals``: None = round to int dollars/count; 0/1/2 = float precision.
_CSR_KPI_INT_FIELDS: tuple[tuple[str, str], ...] = (
    ("aggregateRiskScoreHighCount", "high_risk_items"),
    ("apexPoActionPoCt", "apex_po_action_po_ct"),
    ("criticalShortages", "critical_shortages"),
    ("erpExceptionMsgPoCt", "erp_exception_msg_po_ct"),
    ("latePOCount", "late_pos"),
    ("latePRCount", "late_prs"),
    ("nonCompliantPosCt", "non_compliant_pos_ct"),
    ("openPoCt", "open_po_ct"),
    ("posPlacedInLast30DaysCt", "pos_placed_30d"),
    ("recsCreatedLast30DaysCt", "recs_created_30d"),
    ("shortageItemCount", "shortages"),
    ("shortagesByOrderLines", "shortages_by_order_lines"),
    ("supplierCt", "supplier_ct"),
    ("workbenchOverdueTasksCt", "overdue_tasks"),
)

_CSR_KPI_DEC1_FIELDS: tuple[tuple[str, str], ...] = (
    ("buyerMappingQualityScore", "buyer_mapping_quality"),
    ("clearToBuildPercent", "clear_to_build_pct"),
    ("clearToCommitPercent", "clear_to_commit_pct"),
    ("commitDateCoverage", "commit_date_coverage_pct"),
    ("componentAvailabilityPercent", "component_availability_pct"),
    ("componentAvailabilityPercentProjected", "component_availability_projected_pct"),
    ("dailyActiveBuyersPercent", "daily_active_buyers_pct"),
    ("dailyEngagedBuyersPercent", "daily_engaged_buyers_pct"),
    ("daysCoverage", "days_coverage"),
    ("doiBackwards", "doi_backwards"),
    ("doiForwards", "doi_days"),
    ("inventoryActionFixRateTrailing90Days", "ia_fix_rate_trailing_90d"),
    ("inventoryActionUnableToFixRateTrailing90Days", "ia_unable_to_fix_rate_trailing_90d"),
    ("onOrderDays", "on_order_days"),
    ("supplierCommitDatePercent", "supplier_commit_date_pct"),
    ("weeklyActiveBuyersPercent", "weekly_active_buyers_pct"),
    ("weeklyEngagedIABuyersPercent", "weekly_engaged_ia_buyers_pct"),
    ("weeklyEngagedSuppliersPercent", "weekly_engaged_suppliers_pct"),
)

_CSR_KPI_ROUND_FIELDS: tuple[tuple[str, str], ...] = (
    ("currentFySpend", "current_fy_spend"),
    ("currentWeek52ldnaTarget", "current_week52_ldna_target"),
    ("dailyInventoryUsage", "daily_inventory_usage"),
    ("earlyDeliveriesValue", "early_deliveries_value"),
    ("excessOnOrderObsoleteValue", "excess_on_order_obsolete_value"),
    ("excessOnOrderValuePositive", "excess_on_order_value"),
    ("excessOnhandDemandedValue", "excess_onhand_demanded_value"),
    ("excessOnhandObsoleteValue", "excess_onhand_obsolete_value"),
    ("excessOnhandValuePositive", "excess_on_hand"),
    ("inventoryActionCurrentReportingPeriodOpenValue", "ia_current_period_open_value"),
    ("inventoryActionCurrentReportingPeriodSavings", "savings_current_period"),
    ("inventoryActionOpenValue", "open_ia_value"),
    ("inventoryActionPreviousReportingPeriodSavings", "ia_previous_period_savings"),
    ("manufacturedInventoryValue", "manufactured_inventory_value"),
    ("pastDuePOValue", "past_due_po_value"),
    ("pastDueRequirementValue", "past_due_req_value"),
    ("potentialSavings", "potential_savings"),
    ("potentialToSell", "potential_to_sell"),
    ("previousFySpend", "previous_fy_spend"),
    ("totalOnHandValue", "on_hand_value"),
    ("totalOnOrderValue", "on_order_value"),
)

_CSR_KPI_DEC2_FIELDS: tuple[tuple[str, str], ...] = (
    ("toiBackwards", "toi_backwards"),
    ("toiForwards", "turns_of_inventory"),
)

_CSR_PLAIN_FIELDS: tuple[tuple[str, str], ...] = (
    ("businessUnit", "business_unit"),
    ("customerNdx", "customer_ndx"),
    ("dateCreated", "date_created"),
    ("dateModified", "date_modified"),
    ("division", "division"),
    ("endDate", "end_date"),
    ("factoryNdx", "factory_ndx"),
    ("inventoryActionCurrentReportingPeriod", "ia_current_reporting_period"),
    ("inventoryActionPreviousReportingPeriod", "ia_previous_reporting_period"),
    ("region", "region"),
    ("startDate", "start_date"),
)

# Preferred column order for customer-export markdown / spreadsheet (§13.2).
CSR_MERGED_SITE_EXPORT_COLUMNS: tuple[str, ...] = (
    "factory",
    "site",
    "entity",
    "region",
    "division",
    "business_unit",
    "customer_ndx",
    "factory_ndx",
    "health_score",
    "automated_health_composite",
    "automated_health_override",
    "clear_to_build_pct",
    "clear_to_commit_pct",
    "component_availability_pct",
    "component_availability_projected_pct",
    "shortages",
    "critical_shortages",
    "shortages_by_order_lines",
    "high_risk_items",
    "buyer_mapping_quality",
    "weekly_active_buyers_pct",
    "daily_active_buyers_pct",
    "daily_engaged_buyers_pct",
    "weekly_engaged_ia_buyers_pct",
    "weekly_engaged_suppliers_pct",
    "on_hand_value",
    "on_order_value",
    "excess_on_hand",
    "excess_on_order_value",
    "excess_onhand_demanded_value",
    "excess_onhand_obsolete_value",
    "excess_on_order_obsolete_value",
    "manufactured_inventory_value",
    "early_deliveries_value",
    "past_due_po_value",
    "past_due_req_value",
    "doi_days",
    "doi_backwards",
    "days_coverage",
    "on_order_days",
    "turns_of_inventory",
    "toi_backwards",
    "daily_inventory_usage",
    "late_pos",
    "late_prs",
    "open_po_ct",
    "non_compliant_pos_ct",
    "apex_po_action_po_ct",
    "erp_exception_msg_po_ct",
    "supplier_ct",
    "supplier_commit_date_pct",
    "commit_date_coverage_pct",
    "savings_current_period",
    "open_ia_value",
    "ia_current_period_open_value",
    "ia_previous_period_savings",
    "ia_fix_rate_trailing_90d",
    "ia_unable_to_fix_rate_trailing_90d",
    "ia_current_reporting_period",
    "ia_previous_reporting_period",
    "potential_savings",
    "potential_to_sell",
    "recs_created_30d",
    "pos_placed_30d",
    "overdue_tasks",
    "current_fy_spend",
    "previous_fy_spend",
    "current_week52_ldna_target",
    "start_date",
    "end_date",
    "date_created",
    "date_modified",
    "automated_health_scores",
)

# Short keys for portfolio LLM export §4.2 (token savings). Must be unique (one long name per short key).
CSR_SITE_FIELD_ABBR: dict[str, str] = {
    "factory": "fac",
    "site": "st",
    "entity": "ent",
    "region": "reg",
    "division": "div",
    "business_unit": "bu",
    "customer_ndx": "cndx",
    "factory_ndx": "fndx",
    "health_score": "hs",
    "automated_health_composite": "ahc",
    "automated_health_override": "aho",
    "automated_health_scores": "ahs",
    "clear_to_build_pct": "ctb",
    "clear_to_commit_pct": "ctc",
    "component_availability_pct": "ca",
    "component_availability_projected_pct": "cap",
    "shortages": "sh",
    "critical_shortages": "csh",
    "shortages_by_order_lines": "sbol",
    "high_risk_items": "hri",
    "buyer_mapping_quality": "bmq",
    "weekly_active_buyers_pct": "wab",
    "daily_active_buyers_pct": "dab",
    "daily_engaged_buyers_pct": "deb",
    "weekly_engaged_ia_buyers_pct": "weib",
    "weekly_engaged_suppliers_pct": "wesp",
    "on_hand_value": "ohv",
    "on_order_value": "oov",
    "excess_on_hand": "eoh",
    "excess_on_order_value": "eoov",
    "excess_onhand_demanded_value": "eodv",
    "excess_onhand_obsolete_value": "eoobv",
    "excess_on_order_obsolete_value": "eooobv",
    "manufactured_inventory_value": "miv",
    "early_deliveries_value": "edv",
    "past_due_po_value": "pdpv",
    "past_due_req_value": "pdrv",
    "doi_days": "doi",
    "doi_backwards": "doib",
    "days_coverage": "dcov",
    "on_order_days": "ood",
    "turns_of_inventory": "toi",
    "toi_backwards": "toib",
    "daily_inventory_usage": "diu",
    "late_pos": "lpo",
    "late_prs": "lpr",
    "open_po_ct": "opoc",
    "non_compliant_pos_ct": "ncpoc",
    "apex_po_action_po_ct": "apapoc",
    "erp_exception_msg_po_ct": "eempoc",
    "supplier_ct": "supc",
    "supplier_commit_date_pct": "scdp",
    "commit_date_coverage_pct": "cdcp",
    "savings_current_period": "scp",
    "open_ia_value": "oia",
    "ia_current_period_open_value": "iciapov",
    "ia_previous_period_savings": "iaprs",
    "ia_fix_rate_trailing_90d": "ifr90",
    "ia_unable_to_fix_rate_trailing_90d": "iutfr90",
    "ia_current_reporting_period": "icrp",
    "ia_previous_reporting_period": "iprp",
    "potential_savings": "psav",
    "potential_to_sell": "pts",
    "recs_created_30d": "rc30",
    "pos_placed_30d": "pp30",
    "overdue_tasks": "odt",
    "current_fy_spend": "cfs",
    "previous_fy_spend": "pfs",
    "current_week52_ldna_target": "cw52t",
    "start_date": "sdt",
    "end_date": "edt",
    "date_created": "dcr",
    "date_modified": "dmd",
}

CSR_SITE_FIELD_LEGEND: dict[str, str] = {short: long for long, short in CSR_SITE_FIELD_ABBR.items()}


def abbreviate_csr_site_row(site: dict[str, Any]) -> dict[str, Any]:
    """Rename site-row keys to short forms for portfolio LLM export (unknown keys kept as-is)."""
    return {CSR_SITE_FIELD_ABBR.get(k, k): v for k, v in site.items()}


_SUPPLY_CHAIN_TOTAL_KEYS: tuple[tuple[str, str], ...] = (
    ("on_hand_value", "on_hand"),
    ("on_order_value", "on_order"),
    ("excess_on_hand", "excess_on_hand"),
    ("excess_on_order_value", "excess_on_order"),
    ("excess_onhand_demanded_value", "excess_onhand_demanded"),
    ("excess_onhand_obsolete_value", "excess_onhand_obsolete"),
    ("excess_on_order_obsolete_value", "excess_on_order_obsolete"),
    ("manufactured_inventory_value", "manufactured_inventory"),
    ("early_deliveries_value", "early_deliveries"),
    ("past_due_po_value", "past_due_po"),
    ("past_due_req_value", "past_due_req"),
)

_PLATFORM_VALUE_TOTAL_KEYS: tuple[tuple[str, str], ...] = (
    ("savings_current_period", "total_savings"),
    ("open_ia_value", "total_open_ia_value"),
    ("ia_current_period_open_value", "total_ia_current_period_open_value"),
    ("ia_previous_period_savings", "total_ia_previous_period_savings"),
    ("potential_savings", "total_potential_savings"),
    ("potential_to_sell", "total_potential_to_sell"),
    ("recs_created_30d", "total_recs_created_30d"),
    ("pos_placed_30d", "total_pos_placed_30d"),
    ("overdue_tasks", "total_overdue_tasks"),
    ("current_fy_spend", "total_current_fy_spend"),
    ("previous_fy_spend", "total_previous_fy_spend"),
    ("current_week52_ldna_target", "total_current_week52_ldna_target"),
)


def _build_csr_site_entry(row: dict[str, Any]) -> dict[str, Any]:
    """Map one CS Report ``delta=week`` row to the full per-factory export shape."""
    entry: dict[str, Any] = {
        "factory": row.get("factoryName", "Unknown"),
        "health_score": _health_score_from_row(row),
    }
    _add_site_entity_from_row(row, entry)
    for column, export_key in _CSR_KPI_INT_FIELDS:
        _set_csr_kpi_int(entry, row, column, export_key)
    for column, export_key in _CSR_KPI_DEC1_FIELDS:
        _set_csr_kpi_dec(entry, row, column, export_key, decimals=1)
    for column, export_key in _CSR_KPI_ROUND_FIELDS:
        _set_csr_kpi_dec(entry, row, column, export_key, decimals=None)
    for column, export_key in _CSR_KPI_DEC2_FIELDS:
        _set_csr_kpi_dec(entry, row, column, export_key, decimals=2)
    for column, export_key in _CSR_PLAIN_FIELDS:
        _set_csr_plain(entry, row, column, export_key)
    _add_automated_health_export(row, entry)
    return entry


def _sum_entry_fields(entries: list[dict[str, Any]], mapping: tuple[tuple[str, str], ...]) -> dict[str, Any]:
    totals: dict[str, float] = {}
    for entry in entries:
        for entry_key, total_key in mapping:
            raw = entry.get(entry_key)
            if isinstance(raw, (int, float)):
                totals[total_key] = totals.get(total_key, 0.0) + float(raw)
    return {k: round(v) for k, v in totals.items()}


def _sum_entry_int_fields(entries: list[dict[str, Any]], keys: tuple[str, ...]) -> dict[str, int]:
    totals: dict[str, int] = {}
    for entry in entries:
        for key in keys:
            raw = entry.get(key)
            if isinstance(raw, (int, float)):
                totals[key] = totals.get(key, 0) + int(raw)
    return totals


def csr_merged_site_export_columns(rows: list[dict[str, Any]]) -> list[str]:
    """Return export column order for merged CSR factory rows."""
    seen: set[str] = set()
    for row in rows:
        seen |= {k for k in row.keys() if row.get(k) not in (None, "", [])}
    ordered = [k for k in CSR_MERGED_SITE_EXPORT_COLUMNS if k in seen]
    ordered.extend(sorted(k for k in seen if k not in ordered))
    return ordered


def _csr_site_entries_for_customer(
    customer_name: str,
    *,
    lookup_keys: list[str] | None = None,
) -> tuple[list[dict[str, Any]], str | None, list[str], list[str]]:
    rows, matched_key, tried, matched_csr_customers = _sites_for_customer_lookup(
        customer_name,
        lookup_keys=lookup_keys,
        delta="week",
    )
    if not rows:
        return [], matched_key, tried, matched_csr_customers
    return [_build_csr_site_entry(r) for r in rows], matched_key, tried, matched_csr_customers


def _csr_load_error(customer_name: str, lookup_keys: list[str] | None, tried: list[str]) -> dict[str, Any]:
    return {
        "error": (
            f"No CS Report data for {customer_name!r} "
            f"(lookup_keys={lookup_keys!r}, tried `customer`={tried!r}, delta=week)"
        ),
        "source": "cs_report",
    }


# ── Public API ──


def get_customer_platform_health(
    customer_name: str,
    *,
    lookup_keys: list[str] | None = None,
) -> dict[str, Any]:
    """Health scores, component availability, CTB/CTC, and shortage summary per site."""
    entries, matched_key, tried, matched_csr_customers = _csr_site_entries_for_customer(
        customer_name,
        lookup_keys=lookup_keys,
    )
    if not entries:
        return _csr_load_error(customer_name, lookup_keys, tried)

    display_name = matched_key or customer_name
    health_colors: dict[str, int] = {}
    shortage_totals = _sum_entry_int_fields(entries, ("shortages", "critical_shortages"))

    for entry in entries:
        health = str(entry.get("health_score") or "NONE")
        health_colors[health] = health_colors.get(health, 0) + 1

    qa.check("CS Report platform health loaded")

    return {
        "customer": display_name,
        "source": "cs_report",
        "factory_count": len(entries),
        "csr_customer_names_merged": matched_csr_customers,
        "health_distribution": health_colors,
        "total_shortages": shortage_totals.get("shortages", 0),
        "total_critical_shortages": shortage_totals.get("critical_shortages", 0),
        "sites_sort": "shortages_desc",
        "sites_note": (
            "Per-factory list is sorted by shortages (highest first). When the export "
            "``healthScore`` cell is NONE but ``automatedHealthScores`` is present, health uses "
            "the automated composite (same signal CSR uses for scored sites). Conversion / "
            "project-only rows may remain NONE."
        ),
        "sites": sorted(entries, key=lambda s: s.get("shortages", 0), reverse=True),
    }


def get_customer_supply_chain(
    customer_name: str,
    *,
    lookup_keys: list[str] | None = None,
) -> dict[str, Any]:
    """Inventory values, DOI, excess, and shortage trends per site."""
    entries, matched_key, tried, _matched_csr = _csr_site_entries_for_customer(
        customer_name,
        lookup_keys=lookup_keys,
    )
    if not entries:
        return _csr_load_error(customer_name, lookup_keys, tried)

    display_name = matched_key or customer_name
    qa.check("CS Report supply chain loaded")

    return {
        "customer": display_name,
        "source": "cs_report",
        "factory_count": len(entries),
        "totals": _sum_entry_fields(entries, _SUPPLY_CHAIN_TOTAL_KEYS),
        "sites": sorted(entries, key=lambda s: s.get("on_hand_value", 0), reverse=True),
    }


def get_customer_platform_value(
    customer_name: str,
    *,
    lookup_keys: list[str] | None = None,
) -> dict[str, Any]:
    """ROI metrics: savings, open IA value, recs created, PO activity."""
    entries, matched_key, tried, _matched_csr = _csr_site_entries_for_customer(
        customer_name,
        lookup_keys=lookup_keys,
    )
    if not entries:
        return _csr_load_error(customer_name, lookup_keys, tried)

    display_name = matched_key or customer_name
    totals = _sum_entry_fields(entries, _PLATFORM_VALUE_TOTAL_KEYS)
    qa.check("CS Report platform value loaded")

    return {
        "customer": display_name,
        "source": "cs_report",
        "factory_count": len(entries),
        **totals,
        "sites": sorted(entries, key=lambda s: s.get("savings_current_period", 0), reverse=True),
    }


def load_csr_all_customers_week() -> dict[str, Any]:
    """Build CSR-shaped aggregates by merging ``delta=week`` CS Report rows for every distinct ``customer``.

    Parses the latest XLSX once via :func:`_fetch_latest_report`, then reuses
    :func:`get_customer_platform_health`, :func:`get_customer_supply_chain`, and
    :func:`get_customer_platform_value` per customer. Site rows include ``csr_customer`` for provenance.
    """
    rows = _fetch_latest_report()
    customers = sorted(
        {
            (r.get("customer") or "").strip()
            for r in rows
            if r.get("delta") == "week" and (r.get("customer") or "").strip()
        }
    )
    err: dict[str, Any] = {"error": "No CS Report rows with delta=week", "source": "cs_report"}
    if not customers:
        return {"platform_health": dict(err), "supply_chain": dict(err), "platform_value": dict(err)}

    ph_sites: list[dict[str, Any]] = []
    health_distribution: dict[str, int] = {}
    total_shortages = 0
    total_critical_shortages = 0
    ph_factory_count = 0

    sc_sites: list[dict[str, Any]] = []
    sc_totals: dict[str, float] = defaultdict(float)
    sc_factory_count = 0

    pv_sites: list[dict[str, Any]] = []
    pv_factory_count = 0
    pv_totals: dict[str, float] = defaultdict(float)

    for cn in customers:
        ph = get_customer_platform_health(cn)
        if not ph.get("error"):
            ph_factory_count += int(ph.get("factory_count") or 0)
            total_shortages += int(ph.get("total_shortages") or 0)
            total_critical_shortages += int(ph.get("total_critical_shortages") or 0)
            for hk, hv in (ph.get("health_distribution") or {}).items():
                health_distribution[hk] = health_distribution.get(hk, 0) + int(hv)
            for s in ph.get("sites") or []:
                row = dict(s)
                row["csr_customer"] = cn
                ph_sites.append(row)

        sc = get_customer_supply_chain(cn)
        if not sc.get("error"):
            sc_factory_count += int(sc.get("factory_count") or 0)
            for k, v in (sc.get("totals") or {}).items():
                if isinstance(v, (int, float)):
                    sc_totals[str(k)] += float(v)
            for s in sc.get("sites") or []:
                row = dict(s)
                row["csr_customer"] = cn
                sc_sites.append(row)

        pv = get_customer_platform_value(cn)
        if not pv.get("error"):
            pv_factory_count += int(pv.get("factory_count") or 0)
            for _entry_key, total_key in _PLATFORM_VALUE_TOTAL_KEYS:
                val = pv.get(total_key)
                if isinstance(val, (int, float)):
                    pv_totals[total_key] += float(val)
            for s in pv.get("sites") or []:
                row = dict(s)
                row["csr_customer"] = cn
                pv_sites.append(row)

    label = "All Customers (CS Report aggregate)"
    merged_ph: dict[str, Any] = {
        "customer": label,
        "source": "cs_report",
        "aggregate_scope": "all_customers_week",
        "distinct_csr_customers": len(customers),
        "factory_count": ph_factory_count,
        "health_distribution": health_distribution,
        "total_shortages": total_shortages,
        "total_critical_shortages": total_critical_shortages,
        "sites": sorted(ph_sites, key=lambda s: s.get("shortages", 0), reverse=True),
    }
    merged_sc: dict[str, Any] = {
        "customer": label,
        "source": "cs_report",
        "aggregate_scope": "all_customers_week",
        "distinct_csr_customers": len(customers),
        "factory_count": sc_factory_count,
        "totals": {k: round(v) for k, v in sc_totals.items()},
        "sites": sorted(sc_sites, key=lambda s: s.get("on_hand_value", 0), reverse=True),
    }
    merged_pv: dict[str, Any] = {
        "customer": label,
        "source": "cs_report",
        "aggregate_scope": "all_customers_week",
        "distinct_csr_customers": len(customers),
        "factory_count": pv_factory_count,
        **{k: round(v) for k, v in pv_totals.items()},
        "sites": sorted(pv_sites, key=lambda s: s.get("savings_current_period", 0), reverse=True),
    }
    return {"platform_health": merged_ph, "supply_chain": merged_sc, "platform_value": merged_pv}


def _csr_selection_customer_key(row: dict[str, Any]) -> str:
    ultimate = str(row.get("ultimate_parent") or "").strip()
    if ultimate:
        return ultimate
    return str(row.get("salesforce_label") or row.get("customer") or "").strip()


def selection_lookup_keys_for_llm_export(row: dict[str, Any]) -> list[str]:
    """Ordered CS Report / Jira lookup keys for a top-ARR selection row."""
    keys: list[str] = []

    def add(term: str) -> None:
        s = (term or "").strip()
        if not s:
            return
        if s.lower() in {k.lower() for k in keys}:
            return
        keys.append(s)

    ultimate = str(row.get("ultimate_parent") or "").strip()
    pendo = row.get("pendo_customer_key")
    if ultimate:
        for k in cs_report_lookup_keys_for_account(
            salesforce_label=ultimate,
            pendo_customer_key=pendo,
        ):
            add(k)
    labels = row.get("salesforce_labels")
    if isinstance(labels, list):
        for label in labels:
            for k in cs_report_lookup_keys_for_account(
                salesforce_label=str(label or "").strip(),
                pendo_customer_key=None,
            ):
                add(k)
    sf_label = str(row.get("salesforce_label") or row.get("customer") or "").strip()
    if sf_label:
        for k in cs_report_lookup_keys_for_account(
            salesforce_label=sf_label,
            pendo_customer_key=pendo if not ultimate else None,
        ):
            add(k)
    return keys


def load_csr_for_pendo_customer_export(
    *,
    pendo_prefix: str,
    customer_query: str = "",
) -> dict[str, Any]:
    """Load full CS Report week slices for a single Pendo customer export (all factories).

    Returns platform_health, supply_chain, and platform_value with every factory row —
    no top-N or site sampling (unlike the portfolio LLM export compaction path).
    """
    prefix = (pendo_prefix or "").strip()
    query = (customer_query or "").strip() or prefix
    lookup_keys = cs_report_lookup_keys_for_account(
        salesforce_label=query,
        pendo_customer_key=prefix,
    )
    ph = get_customer_platform_health(prefix, lookup_keys=lookup_keys)
    sc = get_customer_supply_chain(prefix, lookup_keys=lookup_keys)
    pv = get_customer_platform_value(prefix, lookup_keys=lookup_keys)
    matched = None
    merged_csr_names: list[str] = []
    if isinstance(ph, dict) and not ph.get("error"):
        matched = ph.get("customer")
        raw_merged = ph.get("csr_customer_names_merged")
        if isinstance(raw_merged, list):
            merged_csr_names = [str(x) for x in raw_merged if str(x).strip()]
    loaded = not all(
        isinstance(block, dict) and block.get("error")
        for block in (ph, sc, pv)
    )
    block = {
        "scope": "single_customer_pendo_export",
        "delta": "week",
        "pendo_prefix": prefix,
        "customer_query": query,
        "csr_lookup_keys": lookup_keys,
        "csr_matched_lookup_key": matched,
        "csr_customer_names_merged": merged_csr_names,
        "platform_health": ph,
        "supply_chain": sc,
        "platform_value": pv,
    }
    block["merged_sites"] = merge_csr_customer_site_rows(block)
    block["summary"] = csr_customer_summary_from_block(block, factory_count=len(block["merged_sites"]))
    block["csr_loaded"] = loaded and bool(block["merged_sites"])
    return block


_CSR_PENDO_SECTION_KEYS = ("platform_health", "supply_chain", "platform_value")
_CSR_SITE_JOIN_FIELDS = ("factory", "site", "entity")


def _csr_site_join_key(site: dict[str, Any]) -> tuple[str, str, str]:
    return tuple(str(site.get(f) or "").strip().lower() for f in _CSR_SITE_JOIN_FIELDS)  # type: ignore[return-value]


def merge_csr_customer_site_rows(csr_block: dict[str, Any]) -> list[dict[str, Any]]:
    """Union platform_health / supply_chain / platform_value ``sites`` into one row per factory."""
    merged: dict[tuple[str, str, str], dict[str, Any]] = {}
    order: list[tuple[str, str, str]] = []
    for sec_name in _CSR_PENDO_SECTION_KEYS:
        sec = csr_block.get(sec_name)
        if not isinstance(sec, dict) or sec.get("error"):
            continue
        for site in sec.get("sites") or []:
            if not isinstance(site, dict):
                continue
            key = _csr_site_join_key(site)
            if key not in merged:
                merged[key] = {}
                order.append(key)
            merged[key].update(site)
    return [merged[k] for k in order]


def csr_customer_summary_from_block(csr_block: dict[str, Any], *, factory_count: int) -> dict[str, Any]:
    """Customer-level rollups across the three CSR worksheets (no per-factory rows)."""
    summary: dict[str, Any] = {"factory_count": factory_count}
    ph = csr_block.get("platform_health")
    if isinstance(ph, dict) and not ph.get("error"):
        for k in (
            "health_distribution",
            "total_shortages",
            "total_critical_shortages",
            "factory_count",
        ):
            if k in ph and k != "factory_count":
                summary[k] = ph[k]
    sc = csr_block.get("supply_chain")
    if isinstance(sc, dict) and not sc.get("error") and isinstance(sc.get("totals"), dict):
        summary["inventory_totals"] = sc["totals"]
    pv = csr_block.get("platform_value")
    if isinstance(pv, dict) and not pv.get("error"):
        for k in (
            "total_savings",
            "total_open_ia_value",
            "total_ia_current_period_open_value",
            "total_ia_previous_period_savings",
            "total_potential_savings",
            "total_potential_to_sell",
            "total_recs_created_30d",
            "total_pos_placed_30d",
            "total_overdue_tasks",
            "total_current_fy_spend",
            "total_previous_fy_spend",
            "total_current_week52_ldna_target",
        ):
            if k in pv:
                summary[k] = pv[k]
    return summary


def load_csr_top_customers_by_arr(
    selection: list[dict[str, Any]],
) -> dict[str, Any]:
    """Load per-customer CS Report week slices for a ranked ARR selection (LLM export §4).

    *selection* rows should include ``ultimate_parent`` (preferred key), summed ``arr``, and
    constituent ``salesforce_labels`` when grouped by ultimate parent. Legacy rows may use
    ``salesforce_label`` only.
    """
    if not selection:
        err: dict[str, Any] = {"error": "empty selection", "source": "cs_report"}
        return {
            "scope": "top_ultimate_parents_by_arr",
            "top_n": 0,
            "selection_ranked": [],
            "customers": {},
            "platform_health": dict(err),
            "supply_chain": dict(err),
            "platform_value": dict(err),
        }

    customers: dict[str, Any] = {}
    selection_ranked: list[dict[str, Any]] = []

    for row in selection:
        if not isinstance(row, dict):
            continue
        customer_key = _csr_selection_customer_key(row)
        if not customer_key:
            continue
        lookup_keys = selection_lookup_keys_for_llm_export(row)
        ph = get_customer_platform_health(customer_key, lookup_keys=lookup_keys)
        sc = get_customer_supply_chain(customer_key, lookup_keys=lookup_keys)
        pv = get_customer_platform_value(customer_key, lookup_keys=lookup_keys)
        matched = None
        merged_csr_names: list[str] = []
        if isinstance(ph, dict) and not ph.get("error"):
            matched = ph.get("customer")
            raw_merged = ph.get("csr_customer_names_merged")
            if isinstance(raw_merged, list):
                merged_csr_names = [str(x) for x in raw_merged if str(x).strip()]
        customers[customer_key] = {
            "ultimate_parent": row.get("ultimate_parent") or customer_key,
            "salesforce_label": customer_key,
            "salesforce_labels": row.get("salesforce_labels") or [],
            "arr": row.get("arr"),
            "pendo_customer_key": row.get("pendo_customer_key"),
            "csr_lookup_keys": lookup_keys,
            "csr_matched_lookup_key": matched,
            "csr_customer_names_merged": merged_csr_names,
            "csr_lookup_name": matched or (lookup_keys[0] if lookup_keys else customer_key),
            "platform_health": ph,
            "supply_chain": sc,
            "platform_value": pv,
        }
        selection_ranked.append(
            {
                "ultimate_parent": row.get("ultimate_parent") or customer_key,
                "salesforce_label": customer_key,
                "salesforce_labels": row.get("salesforce_labels") or [],
                "arr": row.get("arr"),
                "csr_lookup_keys": lookup_keys,
                "csr_matched_lookup_key": matched,
                "csr_loaded": not all(
                    isinstance(block, dict) and block.get("error")
                    for block in (ph, sc, pv)
                ),
            }
        )

    return {
        "scope": "top_ultimate_parents_by_arr",
        "top_n": len(selection_ranked),
        "aggregate_scope": "top_ultimate_parents_by_arr",
        "note": (
            "Per-customer CS Report (delta=week) for the highest-ARR active Salesforce ultimate "
            "parents (contract rollups summed by ultimate parent — same grouping as "
            "``arr_by_ultimate_parent`` in §3c). Each entry under ``customers`` has "
            "platform_health, supply_chain, and platform_value for that parent group "
            "(not a portfolio-wide merge)."
        ),
        "selection_ranked": selection_ranked,
        "customers": customers,
    }


def cross_validate_with_pendo(customer_name: str, pendo_health: dict[str, Any]) -> None:
    """Compare CS Report data with Pendo data and flag discrepancies via QA."""
    sites = _customer_rows(customer_name, "week")
    if not sites:
        return

    # 1. Site count: Pendo total_sites vs CS Report factory count
    pendo_site_count = pendo_health.get("account", {}).get("total_sites", 0)
    cs_factory_count = len(sites)
    if pendo_site_count == cs_factory_count:
        qa.check("Site count matches between Pendo and CS Report")
    else:
        qa.flag(
            f"Site count differs: Pendo sees {pendo_site_count} sites, CS Report has {cs_factory_count} factories",
            expected=pendo_site_count,
            actual=cs_factory_count,
            sources=("Pendo visitor data", "CS Report factory list"),
            severity="info",
            auto_corrected=False,
        )

    # 2. Engagement: Pendo weekly active rate vs CS Report weekly active buyers
    pendo_rate = pendo_health.get("engagement", {}).get("active_rate_7d", 0)
    cs_buyer_rates = []
    for r in sites:
        wab = _kpi_end(r.get("weeklyActiveBuyersPercent"))
        if wab is not None and wab > 0:
            cs_buyer_rates.append(wab)
    if cs_buyer_rates:
        cs_avg_rate = round(sum(cs_buyer_rates) / len(cs_buyer_rates), 1)
        diff = abs(pendo_rate - cs_avg_rate)
        if diff <= 15:
            qa.check("Pendo and CS Report engagement rates are consistent")
        else:
            qa.flag(
                f"Engagement rate gap: Pendo app login rate {pendo_rate}% vs CS Report buyer engagement {cs_avg_rate}%",
                expected=f"within 15pp",
                actual=f"{diff:.0f}pp difference",
                sources=("Pendo active_rate_7d", "CS Report weeklyActiveBuyersPercent"),
                severity="info" if diff <= 30 else "warning",
                auto_corrected=False,
            )

    # 3. Factory names: check if CS Report factories match Pendo sitenames
    pendo_sites = set()
    for s in pendo_health.get("sites", []):
        if isinstance(s, dict):
            pendo_sites.add(s.get("sitename", "").lower())
    cs_factories = set(r.get("factoryName", "").lower() for r in sites)
    prefix = customer_name.lower()
    pendo_stripped = set(
        s.replace(prefix, "").strip() for s in pendo_sites
    )
    cs_only = cs_factories - pendo_stripped - {""}
    pendo_only = pendo_stripped - cs_factories - {""}
    if cs_only or pendo_only:
        qa.flag(
            f"Site name mismatch: {len(cs_only)} in CS Report only, {len(pendo_only)} in Pendo only",
            expected="matching site lists",
            actual=f"CS-only: {sorted(cs_only)[:5]}, Pendo-only: {sorted(pendo_only)[:5]}",
            sources=("CS Report factoryName", "Pendo sitename"),
            severity="info",
        )
    else:
        qa.check("Factory names consistent between Pendo and CS Report")


def get_csr_section(report: dict[str, Any]) -> dict[str, Any]:
    """Return the CS Report (CSR) subsection from a merged health report.

    Preferred shape: ``report["csr"]`` with keys ``platform_health``, ``supply_chain``,
    ``platform_value``. Legacy top-level ``cs_platform_*`` keys are still supported.
    """
    csr = report.get("csr")
    if isinstance(csr, dict) and csr:
        return csr
    return {
        "platform_health": report.get("cs_platform_health") or {},
        "supply_chain": report.get("cs_supply_chain") or {},
        "platform_value": report.get("cs_platform_value") or {},
    }
